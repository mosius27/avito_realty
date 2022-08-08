# -*- coding: utf-8 -*-

import yaml
import json
import csv
from openpyxl import Workbook, load_workbook
import psycopg2
import sys
sys.path.append('./')
import scripts.other.logger as log
log.Logging()

def read_txt(path: str):
    var = []
    with open(path, 'r', encoding='utf-8') as file:
        for line in file: var.append(line.replace('\n', ''))
    return var

def write_list_in_txt(path: str, var: list):
    with open(path, 'w', encoding='utf-8') as file:
        for element in var: file.write(f'{element}\n')

def create_txt(path):
    with open(path, 'w', encoding='utf-8') as file: pass

def create_csv(path, fieldnames):
    with open(path, 'w', encoding='utf-8', newline='') as file: 
        writer = csv.DictWriter(file, fieldnames=fieldnames.split(';'), delimiter=';', skipinitialspace=True, dialect = 'excel')
        writer.writeheader()

def write_csv(path: str, var: dict, fieldnames: str):
    with open(path, 'a', encoding='utf-8', newline='') as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames.split(';'), delimiter=';', skipinitialspace=True, dialect = 'excel')
        writer.writerow(var)

def create_excel(path: str):
    workbook = Workbook()
    workbook.save(path)

def write_line_excel(path: str, var: str, num_row: int=None):
    workbook = load_workbook(path)
    worksheet = workbook.active
    if num_row != None:
        line = var.split(';')
        col = 1
        for value in line:
            worksheet.cell(row=num_row, column=col, value=value)
            col += 1
    else: worksheet.append(var.split(';'))
    workbook.save(path)

def load_yaml(path: str):
    with open(path, 'r', encoding='utf-8') as file: return yaml.load(file, Loader=yaml.FullLoader)

def read_json(path: str):
    with open(path, 'r', encoding='utf-8') as file: return json.load(file)

def write_json(path: str, var: list or dict):
    with open(path, 'w', encoding='utf8') as file: json.dump(var, file, indent='\t', ensure_ascii=False)

def check_exists_collumn_postgresSQL(settings, var: str):
    try:
        connection = psycopg2.connect(
            host=settings['host'],
            port=settings['port'],
            user=settings['user'],
            password=settings['password'],
            database=settings['db_name']
        )

        connection.autocommit = True

        with connection.cursor() as cursor:
            cursor.execute(
                f"""SELECT EXISTS (SELECT column_name
                    FROM information_schema.columns 
                    WHERE table_name='{settings['table_name']}' AND column_name='{var}');
                    """
            )
            return cursor.fetchone()[0]
    except Exception as _ex:
        log.logger.info("[INFO] Error while working with PostgreSQL", _ex)
    finally:
        if connection:
            connection.close()
            log.logger.info("[INFO] PostgreSQL connection closed")

def create_table_PostgresSQL(settings, var: dict):
    try:
        connection = psycopg2.connect(
            host=settings['host'],
            port=settings['port'],
            user=settings['user'],
            password=settings['password'],
            database=settings['db_name']
        )

        connection.autocommit = True
        with connection.cursor() as cursor:
            cursor.execute(
                f"""CREATE tABLE IF NOT EXISTS {settings['db_name']}(
                    id serial PRIMARY KEY,
                    {var.Дата_публикации} timestamp,
                    {var.Заголовок} text,
                    {var.Тип_недвижимости} text,
                    {var.Описание} text,
                    {var.Цена} integer,
                    {var.Регион} text,
                    {var.Город} text,
                    {var.Адрес} text,
                    {var.Url} text,
                    {var.Изображения} text);
                """
            )
            
            log.logger.info("[INFO] Table created successfully")
    except Exception as _ex:
        log.logger.info("[INFO] Error while working with PostgreSQL", _ex)
    finally:
        if connection:
            connection.close()
            log.logger.info("[INFO] PostgreSQL connection closed")

def insert_table_PostgresSQL(settings, var: dict, params: dict):
    try:
        connection = psycopg2.connect(
            host=settings['host'],
            port=settings['port'],
            user=settings['user'],
            password=settings['password'],
            database=settings['db_name']
        )

        var = dict(var)
        connection.autocommit = True

        for param in params:
            var[param] = params[param]

        keys = values = ''
        for v in var:
            keys += f'{v},'
            values += f"'{var[v]}',"

        if keys[-1] == ',':
            keys = keys[:-1]
        if values[-1] == ',':
            values = values[:-1]

        with connection.cursor() as cursor:
            cursor.execute(f"INSERT INTO {settings['table_name']} ({keys}) VALUES({values});")
            
            log.logger.info("[INFO] Data was successfully inserted")

    except Exception as _ex:
        log.logger.info("[INFO] Error while working with PostgreSQL", _ex)
    finally:
        if connection:
            connection.close()
            log.logger.info("[INFO] PostgreSQL connection closed")

def add_column_PostgresSQL(settings, var: str):
    try:
        connection = psycopg2.connect(
            host=settings['host'],
            port=settings['port'],
            user=settings['user'],
            password=settings['password'],
            database=settings['db_name']
        )
        
        connection.autocommit = True

        with connection.cursor() as cursor:
            cursor.execute(
                f"ALTER TABLE {settings['table_name']} ADD COLUMN {var} text")

    except Exception as _ex:
        log.logger.info("[INFO] Error while working with PostgreSQL", _ex)
    finally:
        if connection:
            connection.close()
            log.logger.info("[INFO] PostgreSQL connection closed")